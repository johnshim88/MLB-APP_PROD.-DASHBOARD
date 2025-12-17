from __future__ import annotations

import os
import re
import traceback
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from openpyxl.worksheet.worksheet import Worksheet

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = BASE_DIR / "★26SS MLB 생산스케쥴_DASHBOARD.xlsx"

FILE_PATH = Path(os.getenv("SUMMARY_EXCEL", str(DEFAULT_WORKBOOK)))
SHEET_NAME = os.getenv("SUMMARY_SHEET", "수량 기준")

# 기본값 (주차를 찾을 수 없을 때 사용)
DEFAULT_WEEK1 = 48
DEFAULT_WEEK2 = 49

# 고정 컬럼: 총 수량
TOTAL_QTY_COL = "C"
TOTAL_QTY_COL_DETAIL = "P"


def _extract_week_from_header(cell_value: Any) -> Optional[int]:
    """셀 값에서 주차 번호를 추출합니다. 'xx주차' 형식을 찾습니다."""
    if cell_value is None:
        return None
    
    # 문자열로 변환
    text = str(cell_value).strip()
    
    # 'xx주차' 패턴 찾기 (예: "49주차", "49 주차" 등)
    match = re.search(r'(\d+)\s*주차', text)
    if match:
        try:
            return int(match.group(1))
        except (ValueError, AttributeError):
            return None
    
    return None


def _find_week_numbers(ws: Worksheet) -> Tuple[int, int]:
    """엑셀 시트의 헤더 행에서 주차 번호를 찾습니다."""
    week_numbers = []
    
    # 헤더 행을 확인 (일반적으로 1-4행)
    for row in range(1, 5):
        # D열부터 L열까지 확인 (첫 번째 주차 그룹)
        for col_letter in ["D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            cell_value = ws[f"{col_letter}{row}"].value
            week_num = _extract_week_from_header(cell_value)
            if week_num is not None and week_num not in week_numbers:
                week_numbers.append(week_num)
                if len(week_numbers) >= 2:
                    break
        if len(week_numbers) >= 2:
            break
    
    # 주차를 찾았으면 정렬 (첫 번째가 금주, 두 번째가 차주)
    if len(week_numbers) >= 2:
        week_numbers.sort()
        return tuple(week_numbers[:2])
    elif len(week_numbers) == 1:
        # 한 개만 찾았으면 차주는 +1
        return (week_numbers[0], week_numbers[0] + 1)
    else:
        # 찾지 못했으면 기본값 사용
        print(f"Warning: Could not find week numbers in header, using defaults: {DEFAULT_WEEK1}, {DEFAULT_WEEK2}")
        return (DEFAULT_WEEK1, DEFAULT_WEEK2)


def _col_num_to_letter(n: int) -> str:
    """열 번호를 엑셀 열 문자로 변환 (1=A, 2=B, ..., 27=AA)"""
    result = ""
    while n > 0:
        n -= 1
        result = chr(65 + (n % 26)) + result
        n //= 26
    return result


def _col_letter_to_num(col: str) -> int:
    """엑셀 열 문자를 열 번호로 변환 (A=1, B=2, ..., AA=27)"""
    result = 0
    for char in col.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def _build_value_columns(week1: int, week2: int, total_qty_col: str = "C", first_week_target_col: str = "D") -> Tuple[Tuple[str, str], ...]:
    """주차 번호에 따라 VALUE_COLUMNS를 동적으로 생성합니다.
    
    Args:
        week1: 첫 번째 주차 번호 (금주)
        week2: 두 번째 주차 번호 (차주)
        total_qty_col: 총 수량 컬럼 (예: "C" 또는 "P")
        first_week_target_col: 첫 번째 주차의 target 컬럼 (예: "D" 또는 "Q")
    """
    # 첫 번째 주차 target 컬럼의 열 번호
    base_col_num = _col_letter_to_num(first_week_target_col)
    
    columns = [
        ("total_qty", total_qty_col),
        (f"target_{week1}", _col_num_to_letter(base_col_num)),      # D 또는 Q
        (f"actual_{week1}", _col_num_to_letter(base_col_num + 1)),  # E 또는 R
        (f"diff_{week1}", _col_num_to_letter(base_col_num + 2)),    # F 또는 S
        (f"target_{week1}_pct", _col_num_to_letter(base_col_num + 3)),  # G 또는 T
        (f"actual_{week1}_pct", _col_num_to_letter(base_col_num + 4)),  # H 또는 U
        (f"target_{week2}", _col_num_to_letter(base_col_num + 5)),      # I 또는 V
        (f"actual_{week2}", _col_num_to_letter(base_col_num + 6)),      # J 또는 W
        (f"target_{week2}_pct", _col_num_to_letter(base_col_num + 7)),  # K 또는 X
        (f"actual_{week2}_pct", _col_num_to_letter(base_col_num + 8)),  # L 또는 Y
    ]
    
    return tuple(columns)


# 동적으로 생성할 예정이므로 None으로 초기화
VALUE_COLUMNS: Optional[Tuple[Tuple[str, str], ...]] = None
DETAIL_VALUE_COLUMNS: Optional[Tuple[Tuple[str, str], ...]] = None
WEEK1: Optional[int] = None
WEEK2: Optional[int] = None

BLOCK_LAYOUT = (
    ("nations", {"rows": range(5, 10), "label_key": "code", "label_col": "B"}),
    ("items", {"rows": range(15, 19), "label_key": "item", "label_col": "B"}),
    (
        "categories",
        {
            "rows": range(24, 80),
            "label_key": "category",
            "label_col": "B",
            "stop_on_blank": True,
            "blank_tolerance": 2,
        },
    ),
    (
        "sub_categories",
        {
            "rows": range(5, 60),
            "label_key": "subcategory",
            "label_col": "O",
            "use_detail_columns": True,  # DETAIL_VALUE_COLUMNS 사용 플래그
            "stop_on_blank": True,
            "blank_tolerance": 3,
        },
    ),
)

app = FastAPI(title="26SS Quantity Summary API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def _extract_block(ws: Worksheet, config: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Generic reader for a contiguous table that shares the same value columns."""

    payload: List[Dict[str, Any]] = []
    blank_streak = 0
    # 동적 컬럼이 설정되지 않은 경우 기본값 사용 (하위 호환성)
    default_columns = _build_value_columns(DEFAULT_WEEK1, DEFAULT_WEEK2, "C", "D")
    
    # value_columns가 config에 명시적으로 설정되어 있으면 그것을 우선 사용
    if "value_columns" in config and config["value_columns"] is not None:
        columns = config["value_columns"]
        print(f"Using value_columns from config: {columns[:2] if columns else 'None'}...")
    else:
        columns = VALUE_COLUMNS or default_columns
        print(f"Using default VALUE_COLUMNS or default_columns")
    
    label_col = config.get("label_col", "B")
    label_key = config.get("label_key", "label")
    rows = config.get("rows", [])
    print(f"Extracting block: label_col={label_col}, label_key={label_key}, rows={list(rows)[:5] if rows else 'empty'}...")

    for row in config["rows"]:
        try:
            label = ws[f"{config['label_col']}{row}"].value

            if label in (None, ""):
                if config.get("stop_on_blank"):
                    blank_streak += 1
                    if blank_streak >= config.get("blank_tolerance", 1):
                        break
                continue

            blank_streak = 0
            entry = {config["label_key"]: label}

            for key, column in columns:
                try:
                    cell_value = ws[f"{column}{row}"].value
                    
                    # #VALUE! 같은 엑셀 오류 문자열 처리
                    if isinstance(cell_value, str) and cell_value.startswith("#"):
                        # 엑셀 오류 값 (#VALUE!, #REF!, #N/A 등)은 None으로 처리
                        entry[key] = None
                        print(f"Warning: Excel error value at {column}{row}: {cell_value}")
                    elif cell_value is None:
                        entry[key] = None
                    elif isinstance(cell_value, (int, float)):
                        # 이미 숫자인 경우 그대로 사용
                        entry[key] = cell_value
                    elif isinstance(cell_value, str):
                        # 문자열 처리
                        cleaned = cell_value.strip().replace(",", "").replace(" ", "")
                        if not cleaned:
                            entry[key] = None
                        elif cleaned.startswith("#"):
                            # 엑셀 오류 값
                            entry[key] = None
                        else:
                            # 숫자로 변환 시도
                            try:
                                # 소수점이 있으면 float, 없으면 int
                                if "." in cleaned:
                                    entry[key] = float(cleaned)
                                else:
                                    entry[key] = int(cleaned)
                            except (ValueError, TypeError):
                                entry[key] = None
                    else:
                        # 기타 타입은 그대로 저장
                        entry[key] = cell_value
                except Exception as e:
                    # 특정 셀 읽기 실패 시 None으로 처리
                    entry[key] = None
                    print(f"Warning: Failed to read {column}{row}: {e}")

            payload.append(entry)
        except Exception as e:
            print(f"Warning: Failed to process row {row}: {e}")
            continue

    return payload


def load_summary(sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """엑셀 파일에서 데이터를 로드하고 주차 정보를 포함하여 반환합니다.
    
    Args:
        sheet_name: 시트 이름 (None이면 기본값 SHEET_NAME 사용)
    """
    global VALUE_COLUMNS, DETAIL_VALUE_COLUMNS, WEEK1, WEEK2
    
    target_sheet = sheet_name or SHEET_NAME
    
    if not FILE_PATH.exists():
        raise FileNotFoundError(f"Excel file not found at {FILE_PATH}")

    workbook = None
    try:
        workbook = openpyxl.load_workbook(FILE_PATH, data_only=True, read_only=True)
    except PermissionError as exc:
        error_msg = (
            f"엑셀 파일에 접근할 수 없습니다. 파일이 다른 프로그램(Excel 등)에서 열려있거나 "
            f"OneDrive 동기화 중일 수 있습니다. 파일을 닫고 다시 시도하세요. 원본 에러: {exc}"
        )
        raise RuntimeError(error_msg) from exc
    except Exception as exc:
        raise RuntimeError(f"Failed to open workbook: {exc}") from exc

    try:
        # 시트 목록 확인 및 로깅
        available_sheets = workbook.sheetnames
        print(f"Available sheets in workbook: {available_sheets}")
        print(f"Looking for sheet: '{target_sheet}'")
        
        if target_sheet not in available_sheets:
            raise RuntimeError(
                f"Worksheet '{target_sheet}' not found. Available sheets: {', '.join(available_sheets)}"
            )
        
        sheet = workbook[target_sheet]
        print(f"Successfully opened sheet: '{target_sheet}'")
        
    except KeyError as exc:
        if workbook:
            workbook.close()
        raise RuntimeError(f"Worksheet '{target_sheet}' not found in workbook") from exc

    try:
        # 헤더에서 주차 정보 추출
        WEEK1, WEEK2 = _find_week_numbers(sheet)
        print(f"Found weeks in Excel: 금주={WEEK1}, 차주={WEEK2}")
        
        # 주차 정보에 따라 동적으로 컬럼 생성
        VALUE_COLUMNS = _build_value_columns(WEEK1, WEEK2, "C", "D")
        DETAIL_VALUE_COLUMNS = _build_value_columns(WEEK1, WEEK2, "P", "Q")
        
        data = {}
        for name, config in BLOCK_LAYOUT:
            try:
                # sub_categories 블록의 경우 DETAIL_VALUE_COLUMNS를 명시적으로 설정
                current_config = config.copy()
                if config.get("use_detail_columns"):
                    current_config["value_columns"] = DETAIL_VALUE_COLUMNS
                    print(f"Block '{name}': Using DETAIL_VALUE_COLUMNS (P~Y columns)")
                    print(f"DETAIL_VALUE_COLUMNS: {DETAIL_VALUE_COLUMNS}")
                elif "value_columns" not in current_config:
                    # 다른 블록들은 VALUE_COLUMNS 사용
                    current_config["value_columns"] = VALUE_COLUMNS
                    print(f"Block '{name}': Using VALUE_COLUMNS (C~L columns)")
                
                extracted = _extract_block(sheet, current_config)
                data[name] = extracted
                print(f"Extracted {len(extracted)} rows for block '{name}'")
                if name == "sub_categories":
                    print(f"=== Sub Categories Debug ===")
                    print(f"Total rows extracted: {len(extracted)}")
                    if len(extracted) > 0:
                        print(f"Sample sub_categories data (first 3 rows):")
                        for i, row in enumerate(extracted[:3]):
                            print(f"  Row {i+1}: {row}")
                    else:
                        print(f"WARNING: No sub_categories data extracted!")
                        print(f"Config used: label_col={current_config.get('label_col')}, rows={list(current_config.get('rows', []))[:10] if current_config.get('rows') else 'None'}")
                        print(f"Value columns used: {current_config.get('value_columns', 'Not set')[:3] if current_config.get('value_columns') else 'Not set'}")
                    print(f"===========================")
            except Exception as e:
                error_msg = f"Failed to extract block '{name}': {str(e)}"
                print(f"Warning: {error_msg}")
                print(traceback.format_exc())
                # 일부 블록 실패해도 다른 블록은 계속 진행
                data[name] = []
        
        # 주차 정보를 포함하여 반환
        result = {
            **data,
            "week_info": {
                "current_week": WEEK1,  # 금주
                "next_week": WEEK2,      # 차주
            },
            "sheet_name": target_sheet,
        }
        
        print(f"Successfully loaded summary data with keys: {list(data.keys())}")
        print(f"Week info: {result['week_info']}")
        return result
    except Exception as e:
        error_msg = f"Unexpected error in load_summary: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        raise RuntimeError(error_msg) from e
    finally:
        if workbook:
            try:
                workbook.close()
            except Exception:
                pass  # close 실패해도 무시


@app.get("/health")
def healthcheck() -> Dict[str, str]:
    return {"status": "ok"}


@app.get("/api/sheets")
def list_sheets() -> Dict[str, Any]:
    """엑셀 파일의 시트 목록을 반환 (디버깅용)"""
    try:
        if not FILE_PATH.exists():
            return {
                "error": f"Excel file not found at {FILE_PATH}",
                "sheets": [],
                "file_exists": False
            }
        
        try:
            workbook = openpyxl.load_workbook(FILE_PATH, read_only=True)
            sheets = workbook.sheetnames
            workbook.close()
            
            return {
                "file_path": str(FILE_PATH),
                "sheets": sheets,
                "current_sheet": SHEET_NAME,
                "sheet_exists": SHEET_NAME in sheets,
                "file_exists": True,
                "file_readable": True
            }
        except PermissionError as exc:
            return {
                "error": f"Permission denied: {str(exc)}",
                "error_type": "PermissionError",
                "suggestion": "엑셀 파일이 다른 프로그램(Excel 등)에서 열려있거나 OneDrive 동기화 중일 수 있습니다. 파일을 닫고 다시 시도하세요.",
                "sheets": [],
                "file_exists": True,
                "file_readable": False
            }
        except Exception as exc:
            return {
                "error": str(exc),
                "error_type": type(exc).__name__,
                "sheets": [],
                "file_exists": True,
                "file_readable": False
            }
    except Exception as exc:
        return {
            "error": str(exc),
            "error_type": type(exc).__name__,
            "sheets": []
        }


@app.get("/api/quantity/debug")
def get_quantity_debug() -> Dict[str, Any]:
    """수량 기준 데이터 디버깅용 - 상세 정보 반환"""
    try:
        data = load_summary()
        
        # nations 배열에서 TOTAL 행 찾기
        nations = data.get("nations", [])
        total_row = None
        for row in nations:
            code = row.get("code", "") or row.get("country", "")
            if str(code).upper() == "TOTAL":
                total_row = row
                break
        
        return {
            "success": True,
            "data_keys": list(data.keys()),
            "nations_count": len(nations),
            "nations_sample": nations[:3] if len(nations) > 0 else [],
            "total_row_found": total_row is not None,
            "total_row": total_row,
            "all_nations_codes": [row.get("code") or row.get("country") or "N/A" for row in nations],
        }
    except Exception as exc:
        return {
            "success": False,
            "error": str(exc),
            "traceback": traceback.format_exc()
        }


@app.get("/api/quantity")
def get_quantity_summary() -> Dict[str, Any]:
    """수량 기준 데이터를 주차 정보와 함께 반환합니다."""
    try:
        return load_summary("수량 기준")
    except (FileNotFoundError, RuntimeError) as exc:
        error_detail = str(exc)
        print(f"Error in /api/quantity: {error_detail}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_detail) from exc
    except Exception as exc:
        error_detail = f"Unexpected error: {str(exc)}"
        print(f"Unexpected error in /api/quantity: {error_detail}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_detail) from exc


@app.get("/api/style-count")
def get_style_count_summary() -> Dict[str, Any]:
    """스타일수 기준 데이터를 주차 정보와 함께 반환합니다."""
    try:
        return load_summary("스타일수 기준")
    except (FileNotFoundError, RuntimeError) as exc:
        error_detail = str(exc)
        print(f"Error in /api/style-count: {error_detail}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_detail) from exc
    except Exception as exc:
        error_detail = f"Unexpected error: {str(exc)}"
        print(f"Unexpected error in /api/style-count: {error_detail}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_detail) from exc

