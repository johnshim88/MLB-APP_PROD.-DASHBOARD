from __future__ import annotations

import os
import re
import json
import traceback
import threading
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, timedelta
from urllib.parse import quote

import openpyxl
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, HTTPException, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.responses import HTMLResponse, Response, FileResponse
from openpyxl.worksheet.worksheet import Worksheet
import secrets

# OneDrive 동기화 모듈 import
try:
    from onedrive_sync import sync_onedrive_file
except ImportError:
    # 로컬 개발 환경에서는 동기화 없이 진행
    def sync_onedrive_file(*args, **kwargs):
        return True

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILENAME = "★26SS MLB 생산스케쥴_DASHBOARD.xlsx"
DEFAULT_WORKBOOK = BASE_DIR / EXCEL_FILENAME

# V2 엑셀 파일명
EXCEL_FILENAME_V2 = "★26SS MLB 생산스케쥴_DASHBOARD_V2.xlsx"
DEFAULT_WORKBOOK_V2 = BASE_DIR / EXCEL_FILENAME_V2

# 환경 변수
FILE_PATH = Path(os.getenv("SUMMARY_EXCEL", str(DEFAULT_WORKBOOK)))
FILE_PATH_V2 = Path(os.getenv("SUMMARY_EXCEL_V2", str(DEFAULT_WORKBOOK_V2)))
SHEET_NAME = os.getenv("SUMMARY_SHEET", "수량 기준")
ONEDRIVE_SHARE_LINK = os.getenv("ONEDRIVE_FILE_URL", "")
ONEDRIVE_SHARE_LINK_V2 = os.getenv("ONEDRIVE_FILE_URL_V2", "")
DASHBOARD_PASSWORD = os.getenv("DASHBOARD_PASSWORD", "MLB123")  # 기본값, 배포 시 변경 필수
SYNC_INTERVAL = int(os.getenv("SYNC_INTERVAL", "3600"))  # 기본 1시간

# 기본값 (주차를 찾을 수 없을 때 사용)
DEFAULT_WEEK1 = 48
DEFAULT_WEEK2 = 49

# 고정 컬럼: 총 수량
TOTAL_QTY_COL = "C"
TOTAL_QTY_COL_DETAIL = "P"

# 비밀번호 인증
security = HTTPBasic()

def verify_password(credentials: HTTPBasicCredentials = Depends(security)) -> bool:
    """비밀번호 인증"""
    correct_password = DASHBOARD_PASSWORD
    is_correct = secrets.compare_digest(credentials.password, correct_password)
    if not is_correct:
        # 디버깅: 잘못된 비밀번호 시도 로그 (보안상 실제 비밀번호는 출력하지 않음)
        print(f"[인증 실패] 사용자명: {credentials.username}, 입력된 비밀번호 길이: {len(credentials.password) if credentials.password else 0}")
        raise HTTPException(
            status_code=401,
            detail="Invalid password",
            headers={"WWW-Authenticate": "Basic"},
        )
    return True


def _extract_week_from_header(cell_value: Any) -> Optional[int]:
    """셀 값에서 주차 번호를 추출합니다. 'xx주차' 형식을 찾습니다."""
    if cell_value is None:
        return None
    
    # 숫자만 있는 경우
    if isinstance(cell_value, (int, float)):
        week_num = int(cell_value)
        if 1 <= week_num <= 60:  # 합리적인 주차 범위
            return week_num
        return None
    
    # 문자열로 변환
    text = str(cell_value).strip()
    
    # 'xx주차' 패턴 찾기 (예: "1주차", "52주차" 등)
    match = re.search(r'(\d+)\s*주차', text)
    if match:
        try:
            week_num = int(match.group(1))
            if 1 <= week_num <= 60:  # 합리적인 주차 범위
                return week_num
            return None
        except (ValueError, AttributeError):
            return None
    
    return None


def _find_week_numbers(ws: Worksheet) -> Tuple[int, int]:
    """엑셀 시트의 헤더 행에서 주차 번호를 찾습니다.
    
    V2 엑셀 파일 구조:
    - D3: "1주차 (금주)" - 실제 주차 번호 (예: 1주차)
    - K3: "2주차 (차주)" - 실제 주차 번호 (예: 2주차)
    """
    week1 = None  # 금주 주차 번호
    week2 = None  # 차주 주차 번호
    
    # D3에서 금주 정보 직접 읽기
    try:
        d3_value = ws["D3"].value
        if d3_value is not None:
            week1 = _extract_week_from_header(d3_value)
            if week1 is not None:
                print(f"금주 정보 발견: D3 = '{d3_value}' -> 주차 {week1}")
            else:
                print(f"Warning: D3에서 주차 번호를 추출할 수 없습니다: '{d3_value}'")
    except Exception as e:
        print(f"Warning: D3 셀 읽기 오류: {e}")
    
    # K3에서 차주 정보 직접 읽기
    try:
        k3_value = ws["K3"].value
        if k3_value is not None:
            week2 = _extract_week_from_header(k3_value)
            if week2 is not None:
                print(f"차주 정보 발견: K3 = '{k3_value}' -> 주차 {week2}")
            else:
                print(f"Warning: K3에서 주차 번호를 추출할 수 없습니다: '{k3_value}'")
    except Exception as e:
        print(f"Warning: K3 셀 읽기 오류: {e}")
    
    # 최종 결과 결정
    if week1 is not None and week2 is not None:
        result = (week1, week2)
        print(f"주차 정보 추출 완료: 금주={result[0]}, 차주={result[1]}")
        return result
    elif week1 is not None:
        result = (week1, week1 + 1)
        print(f"금주만 발견: 금주={result[0]}, 차주={result[1]} (자동 계산)")
        return result
    else:
        # 찾지 못한 경우 기본값 사용
        result = (DEFAULT_WEEK1, DEFAULT_WEEK2)
        print(f"주차 정보를 찾지 못함. 기본값 사용: 금주={result[0]}, 차주={result[1]}")
        return result


def _col_num_to_letter(n: int) -> str:
    """열 번호를 엑셀 열 문자로 변환 (1=A, 2=B, ..., 27=AA)"""
    result = ""
    while n > 0:
        n -= 1
        result = chr(65 + (n % 26)) + result
        n //= 26
    return result


def _col_letter_to_num(col: str) -> int:
    """엑셀 열 문자를 열 번호로 변환 (A=1, B=2, ..., AA=27)"""
    result = 0
    for char in col.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def _build_value_columns(week1: int, week2: int, total_qty_col: str = "C", first_week_target_col: str = "D") -> Tuple[Tuple[str, str], ...]:
    """주차 번호에 따라 VALUE_COLUMNS를 동적으로 생성합니다.
    
    Args:
        week1: 첫 번째 주차 번호 (금주)
        week2: 두 번째 주차 번호 (차주)
        total_qty_col: 총 수량 컬럼 (예: "C" 또는 "P")
        first_week_target_col: 첫 번째 주차의 target 컬럼 (예: "D" 또는 "Q")
    """
    # 첫 번째 주차 target 컬럼의 열 번호
    base_col_num = _col_letter_to_num(first_week_target_col)
    
    columns = [
        ("total_qty", total_qty_col),
        (f"target_{week1}", _col_num_to_letter(base_col_num)),      # D 또는 Q
        (f"actual_{week1}", _col_num_to_letter(base_col_num + 1)),  # E 또는 R
        (f"diff_{week1}", _col_num_to_letter(base_col_num + 2)),    # F 또는 S
        (f"target_{week1}_pct", _col_num_to_letter(base_col_num + 3)),  # G 또는 T
        (f"actual_{week1}_pct", _col_num_to_letter(base_col_num + 4)),  # H 또는 U
        (f"target_{week2}", _col_num_to_letter(base_col_num + 5)),      # I 또는 V
        (f"actual_{week2}", _col_num_to_letter(base_col_num + 6)),      # J 또는 W
        (f"target_{week2}_pct", _col_num_to_letter(base_col_num + 7)),  # K 또는 X
        (f"actual_{week2}_pct", _col_num_to_letter(base_col_num + 8)),  # L 또는 Y
    ]
    
    return tuple(columns)


# 동적으로 생성할 예정이므로 None으로 초기화
VALUE_COLUMNS: Optional[Tuple[Tuple[str, str], ...]] = None
DETAIL_VALUE_COLUMNS: Optional[Tuple[Tuple[str, str], ...]] = None
WEEK1: Optional[int] = None
WEEK2: Optional[int] = None

# 데이터 캐시 시스템
_data_cache: Optional[Dict[str, Any]] = None
_cache_timestamp: Optional[datetime] = None
_cache_lock = threading.Lock()

# V2 데이터 캐시 시스템
_data_cache_v2: Optional[Dict[str, Any]] = None
_cache_timestamp_v2: Optional[datetime] = None
_cache_lock_v2 = threading.Lock()

# 업데이트 시간 설정 (새벽 2시)
UPDATE_HOUR = 2
UPDATE_MINUTE = 0

BLOCK_LAYOUT = (
    ("nations", {"rows": range(5, 10), "label_key": "code", "label_col": "B"}),
    ("items", {"rows": range(15, 19), "label_key": "item", "label_col": "B"}),
    (
        "categories",
        {
            "rows": range(24, 80),
            "label_key": "category",
            "label_col": "B",
            "stop_on_blank": True,
            "blank_tolerance": 2,
        },
    ),
    (
        "sub_categories",
        {
            "rows": range(5, 100),  # 모든 항목 포함을 위해 범위 확대
            "label_key": "subcategory",
            "label_col": "O",
            "use_detail_columns": True,  # DETAIL_VALUE_COLUMNS 사용 플래그
            "stop_on_blank": True,
            "blank_tolerance": 10,  # 더 많은 빈 행 허용
        },
    ),
)

app = FastAPI(title="26SS Quantity Summary API")

# Gzip 압축 미들웨어 추가 (1KB 이상만 압축)
app.add_middleware(GZipMiddleware, minimum_size=1000)

# CORS 설정 - 프론트엔드 도메인만 허용하도록 제한 가능
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 배포 후 프론트엔드 도메인으로 제한 권장
    allow_methods=["*"],
    allow_headers=["*"],
)


def _extract_block(ws: Worksheet, config: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Generic reader for a contiguous table that shares the same value columns.
    최적화: iter_rows()를 사용하여 행 단위 배치 읽기로 성능 향상."""

    payload: List[Dict[str, Any]] = []
    blank_streak = 0
    # 동적 컬럼이 설정되지 않은 경우 기본값 사용 (하위 호환성)
    default_columns = _build_value_columns(DEFAULT_WEEK1, DEFAULT_WEEK2, "C", "D")
    
    # value_columns가 config에 명시적으로 설정되어 있으면 그것을 우선 사용
    if "value_columns" in config and config["value_columns"] is not None:
        columns = config["value_columns"]
    else:
        columns = VALUE_COLUMNS or default_columns
    
    label_col = config.get("label_col", "B")
    label_key = config.get("label_key", "label")
    rows = config.get("rows", [])
    
    if not rows:
        return payload
    
    # 행 범위 계산
    min_row = min(rows)
    max_row = max(rows)
    
    # 열 인덱스 변환 (문자 -> 숫자)
    label_col_num = _col_letter_to_num(label_col)
    column_nums = {key: _col_letter_to_num(col) for key, col in columns}
    
    # 필요한 열 범위 계산 (label_col + 모든 value columns)
    all_cols = [label_col_num] + list(column_nums.values())
    min_col = min(all_cols)
    max_col = max(all_cols)
    
    # iter_rows로 배치 읽기 (values_only=True로 값만 가져오기)
    try:
        row_data = {}
        for row_idx, row_values in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True), start=min_row):
            if row_idx not in rows:
                continue
            
            try:
                # label_col 값 가져오기
                label_idx = label_col_num - min_col
                label = row_values[label_idx] if label_idx < len(row_values) else None
                
                # None이나 빈 문자열 처리
                if label in (None, ""):
                    if config.get("stop_on_blank"):
                        blank_streak += 1
                        if blank_streak >= config.get("blank_tolerance", 1):
                            break
                    continue
                
                blank_streak = 0
                entry = {label_key: label}
                
                # 각 컬럼 값 가져오기
                for key, col_letter in columns:
                    try:
                        col_num = column_nums[key]
                        col_idx = col_num - min_col
                        cell_value = row_values[col_idx] if col_idx < len(row_values) else None
                        
                        # 값 처리 및 최적화
                        if isinstance(cell_value, str) and cell_value.startswith("#"):
                            # 엑셀 오류는 None으로 처리하되 딕셔너리에 추가하지 않음
                            continue
                        elif cell_value is None:
                            # None 값은 추가하지 않음 (데이터 구조 최적화)
                            continue
                        elif isinstance(cell_value, (int, float)):
                            entry[key] = cell_value
                        elif isinstance(cell_value, str):
                            cleaned = cell_value.strip().replace(",", "").replace(" ", "")
                            if cleaned and not cleaned.startswith("#"):
                                try:
                                    entry[key] = float(cleaned) if "." in cleaned else int(cleaned)
                                except (ValueError, TypeError):
                                    continue
                    except Exception:
                        continue
                
                # entry에 데이터가 있으면 추가 (빈 딕셔너리 제외)
                if len(entry) > 1:  # label_key 외에 다른 키가 있으면
                    payload.append(entry)
            except Exception:
                continue
    except Exception:
        # iter_rows 실패 시 기존 방식으로 fallback
        for row in rows:
            try:
                label = ws[f"{label_col}{row}"].value
                if label in (None, ""):
                    if config.get("stop_on_blank"):
                        blank_streak += 1
                        if blank_streak >= config.get("blank_tolerance", 1):
                            break
                    continue
                
                blank_streak = 0
                entry = {label_key: label}
                
                for key, column in columns:
                    try:
                        cell_value = ws[f"{column}{row}"].value
                        if isinstance(cell_value, str) and cell_value.startswith("#"):
                            continue
                        elif cell_value is None:
                            continue
                        elif isinstance(cell_value, (int, float)):
                            entry[key] = cell_value
                        elif isinstance(cell_value, str):
                            cleaned = cell_value.strip().replace(",", "").replace(" ", "")
                            if cleaned and not cleaned.startswith("#"):
                                try:
                                    entry[key] = float(cleaned) if "." in cleaned else int(cleaned)
                                except (ValueError, TypeError):
                                    continue
                    except Exception:
                        continue
                
                if len(entry) > 1:
                    payload.append(entry)
            except Exception:
                continue

    return payload


def ensure_excel_file() -> Path:
    """엑셀 파일이 존재하는지 확인하고, OneDrive에서 동기화합니다."""
    # OneDrive 링크가 설정되어 있고 파일이 없으면 동기화 시도
    if ONEDRIVE_SHARE_LINK and not FILE_PATH.exists():
        sync_onedrive_file(ONEDRIVE_SHARE_LINK, FILE_PATH, sync_interval=SYNC_INTERVAL, force_download=False)
    
    # 파일이 여전히 없으면 에러
    if not FILE_PATH.exists():
        raise FileNotFoundError(f"Excel file not found at {FILE_PATH}. OneDrive sync may have failed.")
    
    return FILE_PATH


def ensure_excel_file_v2() -> Path:
    """V2 엑셀 파일이 존재하는지 확인하고, 필요시 OneDrive에서 동기화합니다."""
    # 기본 경로에서 파일 확인
    if FILE_PATH_V2.exists():
        print(f"V2 파일 발견: {FILE_PATH_V2}")
        return FILE_PATH_V2
    
    # 상위 디렉토리에서도 찾기 (로컬 개발 환경 대응)
    parent_dir = BASE_DIR.parent
    parent_file_path = parent_dir / EXCEL_FILENAME_V2
    if parent_file_path.exists():
        print(f"V2 파일을 상위 디렉토리에서 찾았습니다: {parent_file_path}")
        return parent_file_path
    
    # OneDrive 링크가 설정되어 있고 파일이 없으면 동기화 시도
    if ONEDRIVE_SHARE_LINK_V2:
        print(f"V2 파일이 없습니다. OneDrive에서 동기화 시도 중... (링크: {ONEDRIVE_SHARE_LINK_V2[:50]}...)")
        sync_onedrive_file(ONEDRIVE_SHARE_LINK_V2, FILE_PATH_V2, sync_interval=SYNC_INTERVAL, force_download=True)
        if FILE_PATH_V2.exists():
            print(f"V2 파일 동기화 성공: {FILE_PATH_V2}")
            return FILE_PATH_V2
        else:
            print(f"V2 파일 동기화 실패: {FILE_PATH_V2} 파일이 생성되지 않았습니다.")
    else:
        print(f"ONEDRIVE_SHARE_LINK_V2 환경 변수가 설정되지 않았습니다.")
    
    # V2 파일이 필수이므로 에러 발생
    error_details = [
        f"Excel file V2 not found. V2 file is required.",
        f"",
        f"Checked locations:",
        f"  1. {FILE_PATH_V2}",
        f"  2. {parent_file_path}",
        f"",
        f"Environment variables:",
        f"  - ONEDRIVE_SHARE_LINK_V2: {'SET (check logs for download status)' if ONEDRIVE_SHARE_LINK_V2 else 'NOT SET'}",
        f"  - SUMMARY_EXCEL_V2: {EXCEL_FILENAME_V2}",
        f"",
        f"Please ensure:",
        f"  1. ONEDRIVE_SHARE_LINK_V2 environment variable is set in Render",
        f"  2. The file can be downloaded from OneDrive/Google Drive",
        f"  3. Check Render logs for download errors"
    ]
    raise FileNotFoundError("\n".join(error_details))


def should_update_cache() -> bool:
    """캐시를 업데이트해야 하는지 확인합니다.
    - 캐시가 없으면 업데이트
    - 오전 11시 이후이고 오늘 업데이트하지 않았으면 업데이트
    - 단순 시간 기반 체크만 수행 (파일 변경 시간 체크 제거)
    """
    global _cache_timestamp
    
    now = datetime.now()
    
    # 캐시가 없으면 업데이트
    if _cache_timestamp is None:
        return True
    
    # 오늘 날짜
    today = now.date()
    cache_date = _cache_timestamp.date()
    
    # 오늘 업데이트했으면 스킵 (다음날 11시까지 캐시 사용)
    if cache_date == today:
        return False
    
    # 오늘 업데이트하지 않았고, 오전 11시 이후면 업데이트
    if now.hour >= UPDATE_HOUR:
        return True
    
    return False


def update_cache(force_sync: bool = False) -> None:
    """캐시를 업데이트합니다. 매일 11시에만 실행됩니다.
    
    Args:
        force_sync: True이면 OneDrive에서 강제로 파일을 동기화합니다.
    """
    global _data_cache, _cache_timestamp
    
    # 기존 캐시 백업 (에러 발생 시 복구용)
    old_cache = _data_cache
    old_timestamp = _cache_timestamp
    
    with _cache_lock:
        try:
            # OneDrive 동기화 (파일이 없거나 강제 동기화 요청 시)
            if ONEDRIVE_SHARE_LINK:
                if not FILE_PATH.exists() or force_sync:
                    # 강제 다운로드로 최신 파일 가져오기
                    sync_onedrive_file(ONEDRIVE_SHARE_LINK, FILE_PATH, sync_interval=0, force_download=True)
            
            # 데이터 로드
            quantity_data = load_summary("수량 기준")
            style_count_data = load_summary("스타일수 기준")
            
            # 데이터 유효성 검사
            if not quantity_data or not isinstance(quantity_data, dict):
                raise ValueError("Invalid quantity data")
            if not style_count_data or not isinstance(style_count_data, dict):
                raise ValueError("Invalid style_count data")
            
            # 필수 키 확인
            required_keys = ["nations", "items", "categories", "week_info"]
            if not all(key in quantity_data for key in required_keys):
                raise ValueError("Missing required keys in quantity data")
            if not all(key in style_count_data for key in required_keys):
                raise ValueError("Missing required keys in style_count data")
            
            _data_cache = {
                "quantity": quantity_data,
                "style_count": style_count_data,
            }
            _cache_timestamp = datetime.now()
            
        except Exception as e:
            # 에러 발생 시 기존 캐시 복구
            if old_cache is not None:
                _data_cache = old_cache
                _cache_timestamp = old_timestamp
            else:
                # 기존 캐시도 없으면 None 유지 (다음 요청 시 직접 로드)
                pass


# 백그라운드 업데이트 플래그
_updating_cache = False
_update_lock = threading.Lock()

def get_cached_data(sheet_name: str) -> Dict[str, Any]:
    """캐시된 데이터를 반환합니다. 파일이 변경되었으면 새로 로드합니다."""
    global _data_cache, _updating_cache, _cache_timestamp
    
    # 파일 존재 확인 및 수정 시간 체크
    file_changed = False
    if FILE_PATH.exists() and _cache_timestamp is not None:
        file_mtime_ts = FILE_PATH.stat().st_mtime
        file_mtime = datetime.fromtimestamp(file_mtime_ts)
        # 타임존 없이 비교 (둘 다 naive datetime)
        cache_time_naive = _cache_timestamp.replace(tzinfo=None) if _cache_timestamp.tzinfo else _cache_timestamp
        # 파일이 캐시 이후에 수정되었는지 확인 (1초 여유)
        if file_mtime > cache_time_naive:
            file_changed = True
            print(f"파일이 변경되었습니다. 캐시 시간: {_cache_timestamp}, 파일 수정 시간: {file_mtime}")
    
    # 캐시가 있고 파일이 변경되지 않았으면 캐시 반환
    if _data_cache is not None and not file_changed:
        if sheet_name == "수량 기준":
            cached = _data_cache.get("quantity")
            if cached and isinstance(cached, dict) and len(cached) > 0:
                return cached
        elif sheet_name == "스타일수 기준":
            cached = _data_cache.get("style_count")
            if cached and isinstance(cached, dict) and len(cached) > 0:
                return cached
    
    # 캐시가 없거나 파일이 변경되었으면 새로 로드
    print(f"데이터 새로 로드: sheet_name={sheet_name}, file_changed={file_changed}")
    try:
        data = load_summary(sheet_name)
        # 캐시 업데이트 (파일이 변경된 경우)
        if file_changed:
            print("파일 변경으로 인해 캐시 업데이트 중...")
            if _data_cache is None:
                _data_cache = {}
            if sheet_name == "수량 기준":
                _data_cache["quantity"] = data
            elif sheet_name == "스타일수 기준":
                _data_cache["style_count"] = data
            _cache_timestamp = datetime.now()
        return data
    except Exception as e:
        print(f"데이터 로드 오류: {e}")
        import traceback
        traceback.print_exc()
        # 최소한의 구조라도 반환하여 프론트엔드 에러 방지
        return {
            "nations": [],
            "items": [],
            "categories": [],
            "sub_categories": [],
            "week_info": {
                "current_week": DEFAULT_WEEK1,
                "next_week": DEFAULT_WEEK2,
            },
            "sheet_name": sheet_name,
        }


def get_cached_data_v2(sheet_name: str) -> Dict[str, Any]:
    """V2 캐시된 데이터를 반환합니다. 파일이 변경되었으면 새로 로드합니다."""
    global _data_cache_v2, _cache_timestamp_v2
    
    # 파일 존재 확인 및 수정 시간 체크
    file_changed = False
    if FILE_PATH_V2.exists() and _cache_timestamp_v2 is not None:
        file_mtime_ts = FILE_PATH_V2.stat().st_mtime
        file_mtime = datetime.fromtimestamp(file_mtime_ts)
        # 타임존 없이 비교 (둘 다 naive datetime)
        cache_time_naive = _cache_timestamp_v2.replace(tzinfo=None) if _cache_timestamp_v2.tzinfo else _cache_timestamp_v2
        # 파일이 캐시 이후에 수정되었는지 확인 (1초 여유)
        if file_mtime > cache_time_naive:
            file_changed = True
            print(f"V2 파일이 변경되었습니다. 캐시 시간: {_cache_timestamp_v2}, 파일 수정 시간: {file_mtime}")
    
    # 캐시가 있고 파일이 변경되지 않았으면 캐시 반환
    if _data_cache_v2 is not None and not file_changed:
        if sheet_name == "수량 기준":
            cached = _data_cache_v2.get("quantity")
            if cached and isinstance(cached, dict) and len(cached) > 0:
                return cached
        elif sheet_name == "스타일수 기준":
            cached = _data_cache_v2.get("style_count")
            if cached and isinstance(cached, dict) and len(cached) > 0:
                return cached
    
    # 캐시가 없거나 파일이 변경되었으면 새로 로드
    try:
        data = load_summary_v2(sheet_name)
        
        # 데이터 유효성 검사 - V2 데이터가 비어있으면 에러
        if not data or len(data.get("nations", [])) == 0:
            error_msg = f"V2 데이터가 비어있습니다. V2 파일이 올바르게 로드되었는지 확인하세요. (sheet_name={sheet_name})"
            print(f"Error: {error_msg}")
            raise ValueError(error_msg)
        
        # 캐시 업데이트 (파일이 변경되었거나 캐시가 없는 경우)
        if file_changed or _data_cache_v2 is None:
            if _data_cache_v2 is None:
                _data_cache_v2 = {}
            if sheet_name == "수량 기준":
                _data_cache_v2["quantity"] = data
            elif sheet_name == "스타일수 기준":
                _data_cache_v2["style_count"] = data
            _cache_timestamp_v2 = datetime.now()
        
        return data
    except FileNotFoundError as fnf_e:
        # V2 파일이 없는 경우 - 404로 변환
        print(f"V2 파일을 찾을 수 없음: {fnf_e}")
        import traceback
        traceback.print_exc()
        raise
    except ValueError as ve:
        # 데이터 유효성 검사 실패 - 에러 재발생
        raise
    except Exception as e:
        print(f"V2 데이터 로드 오류: {e}")
        import traceback
        traceback.print_exc()
        
        # V2 파일이 필수이므로 에러 발생 (V1 폴백 제거)
        error_msg = f"V2 데이터를 로드할 수 없습니다: {str(e)}. V2 파일이 올바르게 설정되어 있는지 확인하세요."
        raise RuntimeError(error_msg) from e


def load_summary(sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """엑셀 파일에서 데이터를 로드하고 주차 정보를 포함하여 반환합니다."""
    global VALUE_COLUMNS, DETAIL_VALUE_COLUMNS, WEEK1, WEEK2
    
    target_sheet = sheet_name or SHEET_NAME
    
    # 파일 존재 확인 및 동기화
    excel_path = ensure_excel_file()
    
    workbook = None
    try:
        # read_only=True로 메모리 사용량 최소화
        workbook = openpyxl.load_workbook(excel_path, data_only=True, read_only=True, keep_links=False)
    except PermissionError as exc:
        error_msg = (
            f"엑셀 파일에 접근할 수 없습니다. 파일이 다른 프로그램에서 열려있거나 "
            f"OneDrive 동기화 중일 수 있습니다. 원본 에러: {exc}"
        )
        raise RuntimeError(error_msg) from exc
    except Exception as exc:
        raise RuntimeError(f"Failed to open workbook: {exc}") from exc

    try:
        available_sheets = workbook.sheetnames
        
        if target_sheet not in available_sheets:
            raise RuntimeError(
                f"Worksheet '{target_sheet}' not found. Available sheets: {', '.join(available_sheets)}"
            )
        
        sheet = workbook[target_sheet]
        
    except KeyError as exc:
        if workbook:
            workbook.close()
        raise RuntimeError(f"Worksheet '{target_sheet}' not found in workbook") from exc

    try:
        # 헤더에서 주차 정보 추출
        WEEK1, WEEK2 = _find_week_numbers(sheet)
        
        # 주차 정보에 따라 동적으로 컬럼 생성
        VALUE_COLUMNS = _build_value_columns(WEEK1, WEEK2, "C", "D")
        DETAIL_VALUE_COLUMNS = _build_value_columns(WEEK1, WEEK2, "P", "Q")
        
        data = {}
        for name, config in BLOCK_LAYOUT:
            try:
                current_config = config.copy()
                if config.get("use_detail_columns"):
                    current_config["value_columns"] = DETAIL_VALUE_COLUMNS
                elif "value_columns" not in current_config:
                    current_config["value_columns"] = VALUE_COLUMNS
                
                extracted = _extract_block(sheet, current_config)
                data[name] = extracted
            except Exception as e:
                data[name] = []
        
        result = {
            **data,
            "week_info": {
                "current_week": WEEK1,
                "next_week": WEEK2,
            },
            "sheet_name": target_sheet,
        }
        
        return result
    except Exception as e:
        error_msg = f"Unexpected error in load_summary: {str(e)}"
        raise RuntimeError(error_msg) from e
    finally:
        if workbook:
            try:
                workbook.close()
            except Exception:
                pass


def load_summary_v2(sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """V2 엑셀 파일에서 데이터를 로드하고 주차 정보를 포함하여 반환합니다."""
    global VALUE_COLUMNS, DETAIL_VALUE_COLUMNS, WEEK1, WEEK2
    
    target_sheet = sheet_name or SHEET_NAME
    
    # V2 파일 존재 확인 및 동기화
    excel_path = ensure_excel_file_v2()
    
    if not excel_path.exists():
        raise FileNotFoundError(f"V2 Excel file not found: {excel_path}")
    
    # 파일 유효성 검사 (ZIP 시그니처 확인)
    try:
        with open(excel_path, "rb") as f:
            file_header = f.read(4)
            if file_header[:2] != b'PK':
                # HTML 에러 페이지일 가능성
                f.seek(0)
                preview = f.read(500).decode('utf-8', errors='ignore')
                error_msg = (
                    f"V2 파일이 유효한 Excel 파일이 아닙니다. "
                    f"다운로드된 파일이 HTML 에러 페이지일 수 있습니다. "
                    f"파일 크기: {excel_path.stat().st_size} bytes"
                )
                print(f"ERROR: {error_msg}")
                print(f"File preview: {preview[:200]}")
                raise ValueError(error_msg)
    except ValueError:
        raise
    except Exception as e:
        print(f"Warning: Could not validate file format: {e}")
    
    workbook = None
    try:
        # read_only=True로 메모리 사용량 최소화
        # data_only=True로 먼저 시도 (계산된 값 읽기)
        workbook = openpyxl.load_workbook(excel_path, data_only=True, read_only=True, keep_links=False)
    except PermissionError as exc:
        error_msg = (
            f"엑셀 파일 V2에 접근할 수 없습니다. 파일이 다른 프로그램에서 열려있거나 "
            f"OneDrive 동기화 중일 수 있습니다. 원본 에러: {exc}"
        )
        raise RuntimeError(error_msg) from exc
    except Exception as exc:
        raise RuntimeError(f"Failed to open workbook V2: {exc}") from exc

    try:
        available_sheets = workbook.sheetnames
        if target_sheet not in available_sheets:
            raise RuntimeError(
                f"Worksheet '{target_sheet}' not found in V2 file. Available sheets: {', '.join(available_sheets)}"
            )
        sheet = workbook[target_sheet]
    except KeyError as exc:
        if workbook:
            workbook.close()
        raise RuntimeError(f"Worksheet '{target_sheet}' not found in workbook V2") from exc

    try:
        # 헤더에서 주차 정보 추출
        WEEK1, WEEK2 = _find_week_numbers(sheet)
        
        # 주차 정보에 따라 동적으로 컬럼 생성
        VALUE_COLUMNS = _build_value_columns(WEEK1, WEEK2, "C", "D")
        DETAIL_VALUE_COLUMNS = _build_value_columns(WEEK1, WEEK2, "P", "Q")
        
        data = {}
        for name, config in BLOCK_LAYOUT:
            try:
                current_config = config.copy()
                if config.get("use_detail_columns"):
                    current_config["value_columns"] = DETAIL_VALUE_COLUMNS
                elif "value_columns" not in current_config:
                    current_config["value_columns"] = VALUE_COLUMNS
                
                extracted = _extract_block(sheet, current_config)
                
                # 국가별/아이템별에 누적값 추가 (F열=누적 목표, G열=누적 실제)
                # 차주 데이터도 추가 (M열=차주 목표, N열=차주 실적)
                # 최적화: 행 번호 매핑을 미리 생성하여 반복 접근 최소화
                if name in ["nations", "items"]:
                    label_col = config.get("label_col", "B")
                    label_key = config.get("label_key", "label")
                    rows = config.get("rows", [])
                    
                    # 행 번호 매핑 생성 (한 번만 읽기)
                    row_map = {}
                    for row_idx in rows:
                        try:
                            label_val = sheet[f"{label_col}{row_idx}"].value
                            if label_val:
                                row_map[str(label_val).strip()] = row_idx
                        except:
                            continue
                    
                    # 누적값 컬럼을 배치로 읽기 (F, G, M, N 열)
                    for row_data in extracted:
                        label = str(row_data.get(label_key, "")).strip()
                        row_num = row_map.get(label)
                        
                        if row_num:
                            try:
                                # 배치로 4개 셀 한 번에 읽기
                                cum_target = sheet[f"F{row_num}"].value
                                cum_actual = sheet[f"G{row_num}"].value
                                next_target = sheet[f"M{row_num}"].value
                                next_actual = sheet[f"N{row_num}"].value
                                
                                row_data["target_cumulative"] = float(cum_target) if isinstance(cum_target, (int, float)) else 0
                                row_data["actual_cumulative"] = float(cum_actual) if isinstance(cum_actual, (int, float)) else 0
                                row_data["target_next"] = float(next_target) if isinstance(next_target, (int, float)) else 0
                                row_data["actual_next"] = float(next_actual) if isinstance(next_actual, (int, float)) else 0
                            except Exception:
                                # 실패 시 기본값 설정
                                row_data["target_cumulative"] = 0
                                row_data["actual_cumulative"] = 0
                                row_data["target_next"] = 0
                                row_data["actual_next"] = 0
                
                # 세부 복종별: S열을 직접 읽어서 모든 항목을 순서대로 추출 (최적화: 배치 읽기)
                if name == "sub_categories":
                    sub_categories_data = []
                    start_row = 5
                    end_row = 150
                    seen_indices = set()
                    
                    # 배치 읽기: S, O, W, X, AD, AE 열을 한 번에 읽기
                    cols_to_read = ["S", "O", "W", "X", "AD", "AE"]
                    col_nums = {col: _col_letter_to_num(col) for col in cols_to_read}
                    min_col = min(col_nums.values())
                    max_col = max(col_nums.values())
                    
                    # iter_rows로 배치 읽기
                    for row_idx, row_values in enumerate(sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col, values_only=True), start=start_row):
                        try:
                            # S열 인덱스 (상대 위치 계산)
                            s_idx = col_nums["S"] - min_col
                            s_val = row_values[s_idx] if s_idx < len(row_values) else None
                            
                            if s_val is None:
                                continue
                            
                            # 인덱스 값 처리
                            if isinstance(s_val, str):
                                index_val = s_val.strip()
                            elif isinstance(s_val, (int, float)):
                                index_val = str(int(s_val)) if s_val == int(s_val) else str(s_val)
                            else:
                                index_val = str(s_val).strip()
                            
                            if not index_val or index_val.lower() == "none" or index_val in seen_indices:
                                continue
                            
                            seen_indices.add(index_val)
                            
                            # 다른 열 값 읽기 (상대 위치)
                            o_idx = col_nums["O"] - min_col
                            w_idx = col_nums["W"] - min_col
                            x_idx = col_nums["X"] - min_col
                            ad_idx = col_nums["AD"] - min_col
                            ae_idx = col_nums["AE"] - min_col
                            
                            o_val = row_values[o_idx] if o_idx < len(row_values) else None
                            w_val = row_values[w_idx] if w_idx < len(row_values) else None
                            x_val = row_values[x_idx] if x_idx < len(row_values) else None
                            ad_val = row_values[ad_idx] if ad_idx < len(row_values) else None
                            ae_val = row_values[ae_idx] if ae_idx < len(row_values) else None
                            
                            sub_categories_data.append({
                                "index": index_val,
                                "subcategory": str(o_val).strip() if o_val else "",
                                "target_cumulative": float(w_val) if isinstance(w_val, (int, float)) else 0,
                                "actual_cumulative": float(x_val) if isinstance(x_val, (int, float)) else 0,
                                "target_next": float(ad_val) if isinstance(ad_val, (int, float)) else 0,
                                "actual_next": float(ae_val) if isinstance(ae_val, (int, float)) else 0,
                            })
                        except Exception:
                            continue
                    
                    data[name] = sub_categories_data
                else:
                    data[name] = extracted
            except Exception as e:
                data[name] = []
        
        # V2 특화: 금주 및 차주 summary_cells 값 추출
        summary_cells = {}
        try:
            # 금주 데이터
            d18_val = sheet["D18"].value
            e18_val = sheet["E18"].value
            f18_val = sheet["F18"].value
            g18_val = sheet["G18"].value
            
            summary_cells["D18"] = float(d18_val) if d18_val is not None and isinstance(d18_val, (int, float)) else 0
            summary_cells["E18"] = float(e18_val) if e18_val is not None and isinstance(e18_val, (int, float)) else 0
            summary_cells["F18"] = float(f18_val) if f18_val is not None and isinstance(f18_val, (int, float)) else 0
            summary_cells["G18"] = float(g18_val) if g18_val is not None and isinstance(g18_val, (int, float)) else 0
            
            # 차주 데이터
            k18_val = sheet["K18"].value
            l18_val = sheet["L18"].value
            m18_val = sheet["M18"].value
            n18_val = sheet["N18"].value
            o18_val = sheet["O18"].value
            p18_val = sheet["P18"].value
            
            summary_cells["K18"] = float(k18_val) if k18_val is not None and isinstance(k18_val, (int, float)) else 0
            summary_cells["L18"] = float(l18_val) if l18_val is not None and isinstance(l18_val, (int, float)) else 0
            summary_cells["M18"] = float(m18_val) if m18_val is not None and isinstance(m18_val, (int, float)) else 0
            summary_cells["N18"] = float(n18_val) if n18_val is not None and isinstance(n18_val, (int, float)) else 0
            summary_cells["O18"] = float(o18_val) if o18_val is not None and isinstance(o18_val, (int, float)) else 0
            summary_cells["P18"] = float(p18_val) if p18_val is not None and isinstance(p18_val, (int, float)) else 0
            
        except Exception as e:
            summary_cells = {
                "D18": 0, "E18": 0, "F18": 0, "G18": 0,
                "K18": 0, "L18": 0, "M18": 0, "N18": 0, "O18": 0, "P18": 0
            }
        
        # V2 특화: 협력사 데이터 추출 (AK열=항목명, AO열=누적 목표, AP열=누적 실제)
        # 협력사 데이터는 여러 행에 있을 수 있으므로 행을 순회
        suppliers_data = []
        try:
            # 협력사 데이터가 있는 행 범위 확인 (보통 18행 근처, 더 넓은 범위로 확장)
            # 최적화: 배치 읽기로 변경 (AK, AO, AP, AV, AW 열)
            cols_to_read = ["AK", "AO", "AP", "AV", "AW"]
            col_nums = {col: _col_letter_to_num(col) for col in cols_to_read}
            min_col = min(col_nums.values())
            max_col = max(col_nums.values())
            
            for row_idx, row_values in enumerate(sheet.iter_rows(min_row=5, max_row=50, min_col=min_col, max_col=max_col, values_only=True), start=5):
                try:
                    # AK열 값 (상대 위치)
                    ak_idx = col_nums["AK"] - min_col
                    ak_val = row_values[ak_idx] if ak_idx < len(row_values) else None
                    
                    if ak_val is None:
                        continue
                    
                    # 인덱스 값 처리
                    if isinstance(ak_val, (int, float)):
                        index_val = str(int(ak_val)) if ak_val == int(ak_val) else str(ak_val)
                    else:
                        index_val = str(ak_val).strip() if ak_val else ""
                    
                    if not index_val or index_val.lower() == "none":
                        continue
                    
                    # 14행 특수 처리
                    if row_idx == 14 and ("(주)노브랜드" in index_val or "노브랜드" in index_val):
                        index_val = "(주)노브랜드_WOVEN"
                    
                    # 다른 열 값 읽기 (상대 위치)
                    ao_idx = col_nums["AO"] - min_col
                    ap_idx = col_nums["AP"] - min_col
                    av_idx = col_nums["AV"] - min_col
                    aw_idx = col_nums["AW"] - min_col
                    
                    ao_val = row_values[ao_idx] if ao_idx < len(row_values) else None
                    ap_val = row_values[ap_idx] if ap_idx < len(row_values) else None
                    av_val = row_values[av_idx] if av_idx < len(row_values) else None
                    aw_val = row_values[aw_idx] if aw_idx < len(row_values) else None
                    
                    suppliers_data.append({
                        "name": index_val,
                        "index": index_val,
                        "target_cumulative": float(ao_val) if isinstance(ao_val, (int, float)) else 0,
                        "actual_cumulative": float(ap_val) if isinstance(ap_val, (int, float)) else 0,
                        "target_next": float(av_val) if isinstance(av_val, (int, float)) else 0,
                        "actual_next": float(aw_val) if isinstance(aw_val, (int, float)) else 0,
                        "value": float(ap_val) if isinstance(ap_val, (int, float)) else 0
                    })
                except Exception:
                    continue
            
            # 중복 제거 및 정리 (주)노브랜드 처리 포함
            seen = set()
            unique_suppliers = []
            노브랜드_count = 0
            for supplier in suppliers_data:
                supplier_name = supplier["name"]
                supplier_index = supplier["index"]
                
                # (주)노브랜드 중복 처리
                # 이미 14행은 WOVEN으로 처리되었으므로, 나머지 (주)노브랜드만 카운트
                if "(주)노브랜드" in str(supplier_index) or "(주)노브랜드" in supplier_name:
                    if "(주)노브랜드_WOVEN" not in str(supplier_index) and "(주)노브랜드_WOVEN" not in supplier_name:
                        노브랜드_count += 1
                        # 첫 번째 (주)노브랜드는 그대로, 두 번째부터는 WOVEN으로 변경
                        # 하지만 14행은 이미 WOVEN으로 처리되었으므로 추가 처리 불필요
                
                # index를 기준으로 중복 체크
                index_key = str(supplier_index) if supplier_index else supplier_name
                if index_key not in seen:
                    seen.add(index_key)
                    unique_suppliers.append(supplier)
            suppliers_data = unique_suppliers
                
        except Exception as e:
            suppliers_data = []
        
        result = {
            **data,
            "week_info": {
                "current_week": WEEK1,
                "next_week": WEEK2,
            },
            "sheet_name": target_sheet,
            "summary_cells": summary_cells,  # E18, F18, G18
            "suppliers": suppliers_data,  # 협력사 데이터
        }
        
        # 최종 데이터 유효성 검사
        print(f"DEBUG: 최종 데이터 요약:")
        print(f"  - nations: {len(result.get('nations', []))}개")
        print(f"  - items: {len(result.get('items', []))}개")
        print(f"  - sub_categories: {len(result.get('sub_categories', []))}개")
        print(f"  - suppliers: {len(result.get('suppliers', []))}개")
        print(f"  - summary_cells 샘플: D18={summary_cells.get('D18')}, F18={summary_cells.get('F18')}, G18={summary_cells.get('G18')}")
        
        return result
    except Exception as e:
        error_msg = f"Unexpected error in load_summary_v2: {str(e)}"
        raise RuntimeError(error_msg) from e
    finally:
        if workbook:
            try:
                workbook.close()
            except Exception:
                pass


@app.get("/health")
def healthcheck() -> Dict[str, str]:
    return {"status": "ok"}


@app.get("/api/v2/auth/verify")
def verify_auth(_: bool = Depends(verify_password)) -> Dict[str, str]:
    """인증 검증 전용 엔드포인트 (파일 로드 없이 인증만 확인)"""
    return {"status": "authenticated", "message": "Authentication successful"}


@app.get("/api/password-info")
def get_password_info() -> Dict[str, Any]:
    """비밀번호 설정 정보를 반환합니다 (디버깅용, 보안상 실제 비밀번호는 반환하지 않음)"""
    password_set = os.getenv("DASHBOARD_PASSWORD")
    return {
        "password_source": "environment_variable" if password_set else "default",
        "password_length": len(password_set) if password_set else len("MLB123"),
        "is_default": password_set is None,
        "hint": "MLB123 (기본값)" if password_set is None else "환경 변수에서 설정됨"
    }


@app.get("/api/cache-status")
def get_cache_status(_: bool = Depends(verify_password)) -> Dict[str, Any]:
    """캐시 상태 정보를 반환합니다."""
    file_mtime = None
    file_size = None
    if FILE_PATH.exists():
        stat = FILE_PATH.stat()
        file_mtime = datetime.fromtimestamp(stat.st_mtime).isoformat()
        file_size = stat.st_size
    
    return {
        "cache_timestamp": _cache_timestamp.isoformat() if _cache_timestamp else None,
        "file_modified_time": file_mtime,
        "file_size": file_size,
        "file_exists": FILE_PATH.exists(),
        "cache_age_seconds": (datetime.now() - _cache_timestamp).total_seconds() if _cache_timestamp else None,
        "next_update_time": None,  # 다음 업데이트 시간 계산
        "has_cache": _data_cache is not None,
    }


@app.get("/api/sheets")
def list_sheets() -> Dict[str, Any]:
    """엑셀 파일의 시트 목록을 반환 (디버깅용)"""
    try:
        excel_path = ensure_excel_file()
        
        workbook = openpyxl.load_workbook(excel_path, read_only=True)
        sheets = workbook.sheetnames
        workbook.close()
        
        return {
            "file_path": str(excel_path),
            "sheets": sheets,
            "current_sheet": SHEET_NAME,
            "sheet_exists": SHEET_NAME in sheets,
            "file_exists": True,
            "file_readable": True
        }
    except Exception as exc:
        return {
            "error": str(exc),
            "error_type": type(exc).__name__,
            "sheets": []
        }


@app.get("/api/quantity")
def get_quantity_summary(_: bool = Depends(verify_password)) -> Response:
    """V1 API - V2로 리다이렉트"""
    # V2 엔드포인트로 리다이렉트
    return get_quantity_summary_v2(_)


@app.get("/api/v2/quantity")
def get_quantity_summary_v2(_: bool = Depends(verify_password)) -> Response:
    """V2 수량 기준 데이터를 주차 정보와 함께 반환합니다. (캐시 사용)"""
    try:
        data = get_cached_data_v2("수량 기준")
        
        # 캐시 타임스탬프 및 파일 수정 시간 추가 (V2 버전)
        cache_timestamp = _cache_timestamp_v2.isoformat() if _cache_timestamp_v2 else None
        file_mtime = None
        if FILE_PATH_V2.exists():
            file_mtime = datetime.fromtimestamp(FILE_PATH_V2.stat().st_mtime).isoformat()
        
        # 메타데이터 추가
        data["_meta"] = {
            "cache_timestamp": cache_timestamp,
            "file_modified_time": file_mtime,
            "cache_age_seconds": (datetime.now() - _cache_timestamp_v2).total_seconds() if _cache_timestamp_v2 else None,
        }
        
        # 캐시 헤더 추가 (1시간 캐시, ETag 기반)
        cache_timestamp_str = cache_timestamp or ""
        # ETag 생성 최적화: week_info만 사용하여 해시 계산
        week_info_str = json.dumps(data.get("week_info", {}), sort_keys=True, ensure_ascii=False)
        etag = f'"{hash(week_info_str + cache_timestamp_str)}"'
        headers = {
            "Cache-Control": "public, max-age=3600, stale-while-revalidate=60",
            "ETag": etag,
            "X-Cache-Timestamp": cache_timestamp_str,
            "X-File-Modified": file_mtime or "",
        }
        # JSON 직렬화 최적화 (separators로 공백 제거하여 크기 감소)
        return Response(
            content=json.dumps(data, ensure_ascii=False, separators=(',', ':')),
            media_type="application/json",
            headers=headers
        )
    except FileNotFoundError as fnf_exc:
        # V2 파일이 없는 경우 명확한 에러 메시지
        error_detail = f"V2 Excel file not found. Please check Render environment variables (ONEDRIVE_SHARE_LINK_V2) and logs."
        print(f"ERROR: FileNotFoundError in /api/v2/quantity: {fnf_exc}")
        import traceback
        print("Full traceback:")
        traceback.print_exc()
        raise HTTPException(status_code=404, detail=error_detail) from fnf_exc
    except ValueError as ve:
        # 데이터 유효성 검사 실패
        error_detail = f"Data validation failed: {str(ve)}"
        print(f"ERROR: ValueError in /api/v2/quantity: {ve}")
        import traceback
        print("Full traceback:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=error_detail) from ve
    except RuntimeError as re:
        # V2 데이터 로드 실패
        error_detail = f"Failed to load V2 data: {str(re)}"
        print(f"ERROR: RuntimeError in /api/v2/quantity: {re}")
        import traceback
        print("Full traceback:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=error_detail) from re
    except Exception as exc:
        error_detail = f"Unexpected error: {str(exc)}"
        print(f"ERROR: Unexpected error in /api/v2/quantity: {error_detail}")
        print(f"Error type: {type(exc).__name__}")
        import traceback
        print("Full traceback:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=error_detail) from exc


@app.get("/api/style-count")
def get_style_count_summary(_: bool = Depends(verify_password)) -> Response:
    """V1 API - V2로 리다이렉트"""
    # V2 엔드포인트로 리다이렉트
    return get_style_count_summary_v2(_)


@app.get("/api/v2/style-count")
def get_style_count_summary_v2(_: bool = Depends(verify_password)) -> Response:
    """V2 스타일수 기준 데이터를 주차 정보와 함께 반환합니다. (캐시 사용)"""
    try:
        data = get_cached_data_v2("스타일수 기준")
        
        # 캐시 타임스탬프 및 파일 수정 시간 추가 (V2 버전)
        cache_timestamp = _cache_timestamp_v2.isoformat() if _cache_timestamp_v2 else None
        file_mtime = None
        if FILE_PATH_V2.exists():
            file_mtime = datetime.fromtimestamp(FILE_PATH_V2.stat().st_mtime).isoformat()
        
        # 메타데이터 추가
        data["_meta"] = {
            "cache_timestamp": cache_timestamp,
            "file_modified_time": file_mtime,
            "cache_age_seconds": (datetime.now() - _cache_timestamp_v2).total_seconds() if _cache_timestamp_v2 else None,
        }
        
        # 캐시 헤더 추가 (1시간 캐시, ETag 기반)
        cache_timestamp_str = cache_timestamp or ""
        # ETag 생성 최적화: week_info만 사용하여 해시 계산
        week_info_str = json.dumps(data.get("week_info", {}), sort_keys=True, ensure_ascii=False)
        etag = f'"{hash(week_info_str + cache_timestamp_str)}"'
        headers = {
            "Cache-Control": "public, max-age=3600, stale-while-revalidate=60",
            "ETag": etag,
            "X-Cache-Timestamp": cache_timestamp_str,
            "X-File-Modified": file_mtime or "",
        }
        # JSON 직렬화 최적화 (separators로 공백 제거하여 크기 감소)
        return Response(
            content=json.dumps(data, ensure_ascii=False, separators=(',', ':')),
            media_type="application/json",
            headers=headers
        )
    except FileNotFoundError as fnf_exc:
        # V2 파일이 없는 경우 명확한 에러 메시지
        error_detail = f"V2 Excel file not found. Please check Render environment variables (ONEDRIVE_SHARE_LINK_V2) and logs."
        print(f"FileNotFoundError in /api/v2/style-count: {fnf_exc}")
        print(traceback.format_exc())
        raise HTTPException(status_code=404, detail=error_detail) from fnf_exc
    except Exception as exc:
        error_detail = f"Unexpected error: {str(exc)}"
        print(f"Unexpected error in /api/v2/style-count: {error_detail}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_detail) from exc
        error_detail = f"Unexpected error: {str(exc)}"
        print(f"Unexpected error in /api/v2/style-count: {error_detail}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_detail) from exc


@app.post("/api/refresh")
def refresh_cache(_: bool = Depends(verify_password)) -> Dict[str, Any]:
    """V1 API - V2로 리다이렉트"""
    return refresh_cache_v2(_)


@app.post("/api/v2/refresh")
def refresh_cache_v2(_: bool = Depends(verify_password)) -> Dict[str, Any]:
    """V2 캐시를 강제로 업데이트합니다. OneDrive에서 최신 파일을 가져와서 캐시를 갱신합니다."""
    global _data_cache_v2, _cache_timestamp_v2
    
    try:
        # V2 캐시 강제 초기화 및 재로드
        _data_cache_v2 = None
        _cache_timestamp_v2 = None
        
        # 강제 동기화
        if ONEDRIVE_SHARE_LINK_V2:
            ensure_excel_file_v2()
        
        # 데이터 재로드 (캐시 갱신)
        get_cached_data_v2("수량 기준")
        get_cached_data_v2("스타일수 기준")
        
        return {
            "status": "success",
            "message": "V2 캐시가 성공적으로 업데이트되었습니다.",
            "timestamp": _cache_timestamp_v2.isoformat() if _cache_timestamp_v2 else None
        }
    except Exception as exc:
        error_detail = f"V2 캐시 업데이트 실패: {str(exc)}"
        print(f"Error refreshing V2 cache: {error_detail}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_detail) from exc


@app.get("/api/export/excel")
def export_excel(_: bool = Depends(verify_password)) -> FileResponse:
    """SUMMARY 엑셀 파일을 다운로드합니다."""
    try:
        excel_path = ensure_excel_file()
        
        if not excel_path.exists():
            raise HTTPException(status_code=404, detail="엑셀 파일을 찾을 수 없습니다.")
        
        # 파일명 인코딩 처리 (한글 및 특수문자 파일명 지원)
        filename = excel_path.name
        # RFC 5987 형식으로 인코딩 (UTF-8)
        filename_encoded = quote(filename, safe='')
        # ASCII 호환 파일명 (fallback)
        filename_ascii = filename.encode('ascii', 'ignore').decode('ascii') or "26SS_MLB_DASHBOARD.xlsx"
        
        # Content-Disposition 헤더를 RFC 5987 형식으로 설정
        content_disposition = f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{filename_encoded}"
        
        return FileResponse(
            path=str(excel_path),
            filename=filename_ascii,  # ASCII 파일명 사용
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": content_disposition
            }
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=f"엑셀 파일을 찾을 수 없습니다: {str(exc)}")
    except Exception as exc:
        error_detail = f"엑셀 파일 다운로드 실패: {str(exc)}"
        print(f"Error exporting excel: {error_detail}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_detail) from exc


@app.get("/api/v2/export/excel")
def export_excel_v2(_: bool = Depends(verify_password)) -> FileResponse:
    """V2 SUMMARY 엑셀 파일을 다운로드합니다."""
    try:
        excel_path = ensure_excel_file_v2()
        
        if not excel_path.exists():
            raise HTTPException(status_code=404, detail="V2 엑셀 파일을 찾을 수 없습니다.")
        
        # 파일명 인코딩 처리 (한글 및 특수문자 파일명 지원)
        filename = excel_path.name
        # RFC 5987 형식으로 인코딩 (UTF-8)
        filename_encoded = quote(filename, safe='')
        # ASCII 호환 파일명 (fallback)
        filename_ascii = filename.encode('ascii', 'ignore').decode('ascii') or "26SS_MLB_DASHBOARD_V2.xlsx"
        
        # Content-Disposition 헤더를 RFC 5987 형식으로 설정
        content_disposition = f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{filename_encoded}"
        
        return FileResponse(
            path=str(excel_path),
            filename=filename_ascii,  # ASCII 파일명 사용
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": content_disposition
            }
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=f"V2 엑셀 파일을 찾을 수 없습니다: {str(exc)}")
    except Exception as exc:
        error_detail = f"V2 엑셀 파일 다운로드 실패: {str(exc)}"
        print(f"Error exporting V2 excel: {error_detail}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_detail) from exc


@app.on_event("startup")
async def startup_event():
    """서버 시작 시 초기 캐시 업데이트 및 백그라운드 스레드 시작"""
    # 비밀번호 설정 확인 (디버깅용)
    password_set = os.getenv("DASHBOARD_PASSWORD")
    if password_set:
        print(f"[서버 시작] 환경 변수에서 비밀번호를 사용합니다. (길이: {len(password_set)})")
    else:
        print(f"[서버 시작] 기본 비밀번호 사용: MLB123")
    
    # V2 파일 확인 (서버 시작 시점에 체크, 없어도 서버는 시작)
    try:
        if ONEDRIVE_SHARE_LINK_V2:
            print(f"[서버 시작] V2 파일 확인 중...")
            ensure_excel_file_v2()
            print(f"[서버 시작] V2 파일 확인 완료")
        else:
            print(f"[서버 시작] 경고: ONEDRIVE_SHARE_LINK_V2 환경 변수가 설정되지 않았습니다.")
            if FILE_PATH_V2.exists():
                print(f"[서버 시작] 로컬에 V2 파일이 있습니다: {FILE_PATH_V2}")
            else:
                print(f"[서버 시작] 경고: V2 파일을 찾을 수 없습니다: {FILE_PATH_V2}")
    except Exception as e:
        print(f"[서버 시작] 경고: V2 파일을 확인할 수 없습니다 (서버는 계속 시작됩니다): {e}")
    
    # 초기 캐시 업데이트 (캐시가 없으면)
    if _data_cache is None:
        update_cache()
    
    # 백그라운드 스레드에서 매일 11시에 업데이트 체크
    def background_update_check():
        """백그라운드에서 매일 11시에 업데이트를 수행합니다."""
        global _updating_cache
        
        while True:
            now = datetime.now()
            
            # 다음 업데이트 시간 계산 (오늘 새벽 2시 또는 내일 새벽 2시)
            if now.hour < UPDATE_HOUR:
                # 오늘 새벽 2시까지 대기
                next_update = now.replace(hour=UPDATE_HOUR, minute=UPDATE_MINUTE, second=0, microsecond=0)
            else:
                # 내일 새벽 2시까지 대기
                next_update = (now + timedelta(days=1)).replace(hour=UPDATE_HOUR, minute=UPDATE_MINUTE, second=0, microsecond=0)
            
            wait_seconds = (next_update - now).total_seconds()
            
            # 다음 업데이트 시간까지 대기 (최대 1시간씩 체크)
            while wait_seconds > 0:
                sleep_time = min(wait_seconds, 3600)  # 최대 1시간씩 체크
                time.sleep(sleep_time)
                wait_seconds -= sleep_time
                
                # 중간에 업데이트 시간이 되었는지 확인
                now = datetime.now()
                if now.hour >= UPDATE_HOUR and should_update_cache():
                    break
            
            # 업데이트 시간이 되었는지 확인
            try:
                with _update_lock:
                    if should_update_cache() and not _updating_cache:
                        _updating_cache = True
                        # V2 캐시 강제 업데이트
                        try:
                            if ONEDRIVE_SHARE_LINK_V2:
                                ensure_excel_file_v2()
                            # V2 데이터 재로드
                            get_cached_data_v2("수량 기준")
                            get_cached_data_v2("스타일수 기준")
                            print(f"V2 cache updated at {datetime.now()}")
                        except Exception as e:
                            print(f"Error updating V2 cache: {e}")
                        _updating_cache = False
            except Exception:
                _updating_cache = False
                pass  # 에러 발생해도 계속 실행
    
    thread = threading.Thread(target=background_update_check, daemon=True)
    thread.start()

